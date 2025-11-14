"""
Utility functions for customer management.
"""
import qrcode
from io import BytesIO
from django.core.mail import EmailMessage
from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def generate_qr_code(customer_id):
    """
    Generate QR code for customer ID.
    Returns a BytesIO buffer containing the PNG image.
    """
    # Create QR code instance
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    
    # Add customer ID data
    qr.add_data(customer_id)
    qr.make(fit=True)
    
    # Create image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to BytesIO
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    return buffer


def send_customer_welcome_email(customer):
    """
    Send welcome email to customer with their unique ID and QR code.
    Generates QR code on-the-fly without saving to database.
    """
    print(f"=== Attempting to send email to {customer.email} ===")
    subject = f'Welcome! Your Customer ID: {customer.customer_id}'
    
    # Create email body
    message = f"""
    Dear {customer.name},

    Thank you for registering with us!

    Your unique Customer ID is: {customer.customer_id}

    Please save this ID or use the attached QR code for future reference.

    Best regards,
    Exhibition Team
    """
    
    # Create email
    email = EmailMessage(
        subject=subject,
        body=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[customer.email],
    )
    
    # Generate QR code on-the-fly and attach to email
    qr_code_buffer = generate_qr_code(customer.customer_id)
    email.attach(
        filename=f'qr_{customer.customer_id}.png',
        content=qr_code_buffer.read(),
        mimetype='image/png'
    )
    
    # Send email
    try:
        print(f"=== Sending email with QR code to {customer.email} ===")
        result = email.send()
        print(f"=== Email send result: {result} ===")
        return True
    except Exception as e:
        print(f"=== ERROR sending email: {e} ===")
        import traceback
        traceback.print_exc()
        return False


def export_customers_to_excel(customers_queryset):
    """
    Export customers with their billing information to Excel format.
    Returns a BytesIO object containing the Excel file.
    """
    # Create workbook and worksheets
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create Summary sheet
    ws_summary = wb.create_sheet("Customers Summary")
    
    # Create Detailed Bills sheet
    ws_details = wb.create_sheet("Detailed Bills")
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === SUMMARY SHEET ===
    summary_headers = [
        'Customer ID', 'Name', 'Email', 'Phone', 
        'Number of Bills', 'Total Amount', 'Email Sent', 
        'Created At', 'Updated At'
    ]
    
    # Write summary headers
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write summary data
    row_num = 2
    for customer in customers_queryset:
        ws_summary.cell(row=row_num, column=1, value=customer.customer_id).border = border
        ws_summary.cell(row=row_num, column=2, value=customer.name).border = border
        ws_summary.cell(row=row_num, column=3, value=customer.email).border = border
        ws_summary.cell(row=row_num, column=4, value=customer.phone).border = border
        ws_summary.cell(row=row_num, column=5, value=customer.get_bill_count()).border = border
        
        # Format total amount
        total_cell = ws_summary.cell(row=row_num, column=6, value=float(customer.get_total_bills()))
        total_cell.number_format = '$#,##0.00'
        total_cell.border = border
        
        ws_summary.cell(row=row_num, column=7, value='Yes' if customer.email_sent else 'No').border = border
        ws_summary.cell(row=row_num, column=8, value=customer.created_at.strftime('%Y-%m-%d %H:%M')).border = border
        ws_summary.cell(row=row_num, column=9, value=customer.updated_at.strftime('%Y-%m-%d %H:%M')).border = border
        
        row_num += 1
    
    # Auto-adjust column widths for summary
    for col in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18
    
    # === DETAILED BILLS SHEET ===
    details_headers = [
        'Customer ID', 'Customer Name', 'Customer Email', 
        'Bill Amount', 'Bill Description', 'Bill Date', 'Created By'
    ]
    
    # Write details headers
    for col_num, header in enumerate(details_headers, 1):
        cell = ws_details.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Write detailed bills data
    row_num = 2
    for customer in customers_queryset:
        bills = customer.bills.all().order_by('-created_at')
        
        if bills.exists():
            for bill in bills:
                ws_details.cell(row=row_num, column=1, value=customer.customer_id).border = border
                ws_details.cell(row=row_num, column=2, value=customer.name).border = border
                ws_details.cell(row=row_num, column=3, value=customer.email).border = border
                
                # Format bill amount
                amount_cell = ws_details.cell(row=row_num, column=4, value=float(bill.amount))
                amount_cell.number_format = '$#,##0.00'
                amount_cell.border = border
                
                ws_details.cell(row=row_num, column=5, value=bill.description or 'N/A').border = border
                ws_details.cell(row=row_num, column=6, value=bill.created_at.strftime('%Y-%m-%d %H:%M')).border = border
                ws_details.cell(row=row_num, column=7, value=bill.created_by or 'N/A').border = border
                
                row_num += 1
        else:
            # Customer with no bills
            ws_details.cell(row=row_num, column=1, value=customer.customer_id).border = border
            ws_details.cell(row=row_num, column=2, value=customer.name).border = border
            ws_details.cell(row=row_num, column=3, value=customer.email).border = border
            ws_details.cell(row=row_num, column=4, value='No bills').border = border
            ws_details.cell(row=row_num, column=5, value='').border = border
            ws_details.cell(row=row_num, column=6, value='').border = border
            ws_details.cell(row=row_num, column=7, value='').border = border
            
            row_num += 1
    
    # Auto-adjust column widths for details
    for col in range(1, len(details_headers) + 1):
        if col == 5:  # Description column wider
            ws_details.column_dimensions[get_column_letter(col)].width = 40
        else:
            ws_details.column_dimensions[get_column_letter(col)].width = 18
    
    # Freeze top row in both sheets
    ws_summary.freeze_panes = 'A2'
    ws_details.freeze_panes = 'A2'
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file.getvalue()

